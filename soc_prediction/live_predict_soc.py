from fastapi import FastAPI, Request
import uvicorn
import numpy as np
import joblib
from datetime import datetime
from openpyxl import Workbook, load_workbook
import tflite_runtime.interpreter as tflite


interpreter = tflite.Interpreter(model_path="soc_model.tflite")
interpreter.allocate_tensors()

scaler_X = joblib.load("scaler_X.pkl")

input_details = interpreter.get_input_details()
output_details = interpreter.get_output_details()

app = FastAPI(title="Battery SoC Predictor (Voltage + Temperature)")

# ======================================================
# EXCEL LOG
# ======================================================
start_time = datetime.now().strftime("%Y%m%d_%H%M%S")
EXCEL_FILE = f"soc_log_{start_time}.xlsx"

wb = Workbook()
ws = wb.active
ws.append([
    "Timestamp",
    "Voltage(V)",
    "Current(A)",
    "Temperature(C)",
    "Predicted_SoC(%)"
])
wb.save(EXCEL_FILE)

# ======================================================
# VOLTAGE VALIDATION
# ======================================================
last_valid_voltage = None

def is_valid_voltage(new_v):
    global last_valid_voltage

    if new_v < 2.5 or new_v > 4.35:
        return False

    if last_valid_voltage is not None:
        if abs(new_v - last_valid_voltage) > 1.2:
            return False

    return True



def predict_soc(voltage, temperature):

    X = np.array([[voltage, temperature]], dtype=np.float32)
    X_scaled = scaler_X.transform(X).astype(np.float32)

    interpreter.set_tensor(input_details[0]['index'], X_scaled)
    interpreter.invoke()
    output = interpreter.get_tensor(output_details[0]['index'])

    soc = float(output[0][0] * 100)
    return float(np.clip(soc, 0, 100))


# ======================================================
# LOGGING
# ======================================================
def log_to_excel(timestamp, voltage, current, temperature, pred):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([
        timestamp,
        round(voltage, 3),
        round(current, 3),
        round(temperature, 2),
        round(pred, 2)
    ])
    wb.save(EXCEL_FILE)


# ======================================================
# API ENDPOINT
# ======================================================
@app.post("/data")
async def receive_data(request: Request):

    global last_valid_voltage

    data = await request.json()

    voltage = float(data["voltage"])
    current = float(data["current"])   
    temperature = float(data["temperature"])

    if not is_valid_voltage(voltage):
        return {"status": "ignored"}

    last_valid_voltage = voltage

    pred_soc = predict_soc(voltage, temperature)

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    log_to_excel(timestamp, voltage, current, temperature, pred_soc)

    print("\n===================================================")
    print(f"📥 DATA RECEIVED @ {timestamp}")
    print(f"Voltage      : {voltage:.3f} V")
    print(f"Current      : {current:.3f} A")
    print(f"Temperature  : {temperature:.2f} °C")
    print(f"Predicted SoC: {pred_soc:.2f} %")
    print("===================================================\n")

    return {
        "predicted_soc": round(pred_soc, 2),
        "current": round(current, 3)
    }


# ======================================================
# RUN SERVER
# ======================================================
if __name__ == "__main__":
    uvicorn.run(
        "live_predict_soc:app",
        host="0.0.0.0",
        port=3000,
        reload=True
    )
